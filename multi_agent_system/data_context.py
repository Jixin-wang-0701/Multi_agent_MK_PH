from __future__ import annotations

from datetime import datetime
from pathlib import Path
import json
import re
import zipfile
from xml.etree import ElementTree as ET

from .config import DATA_FILES
from .r_support import diagnostics_markdown, summarize_rds_with_r


MAX_DOCX_CHARS = 12000
MAX_CELL_CHARS = 120


def _file_overview(path: Path) -> str:
    stat = path.stat()
    size_mb = stat.st_size / (1024 * 1024)
    modified = datetime.fromtimestamp(stat.st_mtime).isoformat(timespec="seconds")
    return f"- Path: {path.name}\n- Size: {size_mb:.2f} MB\n- Modified: {modified}"


def _clean_text(value: object, max_chars: int = MAX_CELL_CHARS) -> str:
    text = "" if value is None else str(value)
    text = re.sub(r"\s+", " ", text).strip()
    if len(text) > max_chars:
        return text[: max_chars - 3] + "..."
    return text


def summarize_docx(path: Path, max_chars: int = MAX_DOCX_CHARS) -> str:
    try:
        from docx import Document  # type: ignore

        document = Document(str(path))
        paragraphs = [_clean_text(p.text, 1000) for p in document.paragraphs]
        paragraphs = [p for p in paragraphs if p]
        text = "\n".join(paragraphs)
    except Exception:
        text = _extract_docx_text_stdlib(path)

    if not text:
        return "DOCX text extraction produced no visible paragraphs."
    suffix = ""
    if len(text) > max_chars:
        suffix = f"\n\n[Truncated from {len(text)} characters.]"
        text = text[:max_chars]
    return text + suffix


def _extract_docx_text_stdlib(path: Path) -> str:
    with zipfile.ZipFile(path) as archive:
        xml = archive.read("word/document.xml")
    root = ET.fromstring(xml)
    namespace = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    paragraphs: list[str] = []
    for paragraph in root.findall(".//w:p", namespace):
        runs = [node.text or "" for node in paragraph.findall(".//w:t", namespace)]
        text = _clean_text("".join(runs), 1000)
        if text:
            paragraphs.append(text)
    return "\n".join(paragraphs)


def summarize_xlsx(path: Path, max_sheets: int = 8, max_rows: int = 6, max_cols: int = 8) -> str:
    try:
        import openpyxl  # type: ignore

        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        lines = [f"Workbook sheets: {', '.join(workbook.sheetnames)}"]
        for sheet in workbook.worksheets[:max_sheets]:
            lines.append(f"\nSheet: {sheet.title}")
            lines.append(f"- Dimensions: {sheet.max_row} rows x {sheet.max_column} columns")
            preview: list[list[str]] = []
            for row in sheet.iter_rows(max_row=max_rows, max_col=max_cols, values_only=True):
                preview.append([_clean_text(cell) for cell in row])
            lines.append("- Preview:")
            lines.append(_format_table_preview(preview))
        if len(workbook.worksheets) > max_sheets:
            lines.append(f"\n[Only first {max_sheets} sheets summarized.]")
        workbook.close()
        return "\n".join(lines)
    except Exception as exc:  # noqa: BLE001 - return a readable fallback.
        return f"Unable to summarize workbook with openpyxl: {exc}"


def _format_table_preview(rows: list[list[str]]) -> str:
    if not rows:
        return "  [empty preview]"
    return "\n".join("  - " + " | ".join(cell or "" for cell in row) for row in rows)


def summarize_rds(path: Path, root: Path) -> str:
    script_path = root / "multi_agent_system" / "r_scripts" / "summarize_seurat_rds.R"
    return summarize_rds_with_r(path, script_path)


def build_data_context(root: Path) -> str:
    sections = [
        "# User-provided data context",
        "",
        "This context is generated locally from available files. Agents must distinguish direct evidence "
        "from extracted summaries, indirect inference, literature knowledge, and speculation.",
    ]

    for label, filename in DATA_FILES.items():
        path = root / filename
        sections.extend(["", f"## {label}: {filename}"])
        if not path.exists():
            sections.append("Status: missing.")
            continue
        sections.append(_file_overview(path))
        suffix = path.suffix.lower()
        if suffix == ".docx":
            sections.extend(["", "Extracted text excerpt:", summarize_docx(path)])
        elif suffix == ".xlsx":
            sections.extend(["", "Workbook summary:", summarize_xlsx(path)])
        elif suffix == ".rds":
            sections.extend(["", summarize_rds(path, root)])
        else:
            sections.append("No extractor configured for this file type.")

    return "\n".join(sections).strip() + "\n"


def write_data_context(root: Path, output_dir: Path) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    context = build_data_context(root)
    path = output_dir / "data_context.md"
    path.write_text(context, encoding="utf-8")

    manifest = {
        "generated": datetime.now().isoformat(timespec="seconds"),
        "files": {label: filename for label, filename in DATA_FILES.items()},
        "context_file": str(path),
    }
    (output_dir / "data_context_manifest.json").write_text(
        json.dumps(manifest, indent=2),
        encoding="utf-8",
    )
    (output_dir / "r_environment.md").write_text(diagnostics_markdown(), encoding="utf-8")
    return path
