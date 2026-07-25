from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
import csv
import json
import math
import re
import statistics
import subprocess
import time
from typing import Any
from urllib import parse, request

from .r_support import select_rscript


METABOLITE_FILES = {
    "mk_metabolomics": "sFig6A Raw data.xlsx",
    "ph_control_metabolomics": "Figure6D+F raw data.xlsx",
}

KEGG_BASE = "https://rest.kegg.jp"
NCBI_BASE = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils"
MAX_PUBMED_GENES = 8
STRUCTURED_BRIDGE_READINESS_MIN = 25.0
STRUCTURED_BRIDGE_RETRIEVE_PER_AXIS = 5
STRUCTURED_BRIDGE_DISPLAY_CANDIDATES_PER_CHAIN = 6


@dataclass(frozen=True)
class DifferentialMetabolite:
    source: str
    metabolite: str
    comparison: str
    direction: str
    log2fc: float
    control_mean: float
    disease_mean: float
    p_value: float | None = None
    fdr: float | None = None
    priority_score: float = 0.0


@dataclass(frozen=True)
class KeggMapping:
    metabolite: str
    compound_ids: list[str]
    compound_names: list[str]
    pathways: list[str]
    enzymes: list[str]
    mouse_genes: list[dict[str, str]]
    neighbor_mouse_genes: list[dict[str, str]]
    neighbor_expanded: bool
    status: str


GENERIC_KEGG_PATHWAYS = {
    "path:map01100",  # Metabolic pathways
    "path:map01110",  # Biosynthesis of secondary metabolites
    "path:map01120",  # Microbial metabolism in diverse environments
    "path:map01200",  # Carbon metabolism
    "path:map01210",  # 2-Oxocarboxylic acid metabolism
    "path:map01212",  # Fatty acid metabolism
    "path:map01230",  # Biosynthesis of amino acids
}

PATHWAY_NEIGHBOR_SCHEMA_VERSION = 2

STRONG_MECHANISM_TAGS = {
    "polyamine metabolism",
    "s-adenosylmethionine metabolism",
    "methionine salvage",
    "cysteine/methionine metabolism",
    "arginine/proline metabolism",
    "glutathione/redox metabolism",
    "tryptophan metabolism",
    "retinoid metabolism",
    "one-carbon metabolism",
}

CANDIDATE_BRIDGE_TERMS = (
    "T cell",
    "T cell differentiation",
    "CD4 T cell",
    "T helper",
    "Th1",
    "Th2",
    "Th17",
    "T helper 17",
    "IL-17",
    "IL-17A",
    "Treg",
    "regulatory T cell",
    "macrophage",
    "macrophage polarization",
    "M1 macrophage",
    "M2 macrophage",
    "monocyte",
    "monocyte recruitment",
    "neutrophil",
    "NETosis",
    "neutrophil recruitment",
    "endothelial activation",
    "endothelial dysfunction",
    "smooth muscle proliferation",
    "smooth muscle migration",
    "fibroblast activation",
    "extracellular matrix",
    "extracellular vesicle",
    "exosome",
    "stromal remodeling",
)

VASCULAR_BRIDGE_TERMS = (
    "pulmonary hypertension",
    "pulmonary arterial hypertension",
    "pulmonary vascular remodeling",
    "vascular remodeling",
    "smooth muscle proliferation",
    "endothelial dysfunction",
    "fibroblast activation",
    "vascular inflammation",
)

MK_BRIDGE_SOURCE_TERMS = (
    "megakaryocyte",
    "megakaryocytes",
    "megakaryocyte-derived",
    "platelet-derived",
    "platelet microparticle",
    "platelet extracellular vesicle",
    "megakaryocyte extracellular vesicle",
    "platelet factor 4",
)

STRUCTURED_DISEASE_BRIDGE_TERMS = (
    "pulmonary hypertension",
    "pulmonary arterial hypertension",
    "pulmonary vascular remodeling",
    "pulmonary vascular remodelling",
    "pulmonary artery remodeling",
)

STRUCTURED_BRIDGE_PROFILES = (
    {
        "key": "generic_cd4_t_helper",
        "label": "generic CD4/T helper",
        "terms": ("CD4 T cell", "T helper", "T cell differentiation", "T cell activation"),
        "reason": "broad T-cell differentiation bridge audit",
    },
    {
        "key": "th1_ifng",
        "label": "Th1/IFN-gamma",
        "terms": ("Th1", "T helper 1", "IFN-gamma", "IFNG", "TBX21"),
        "reason": "T-cell subset bridge audit",
    },
    {
        "key": "th2_il4",
        "label": "Th2/IL-4",
        "terms": ("Th2", "T helper 2", "IL-4", "IL-13", "GATA3"),
        "reason": "T-cell subset bridge audit",
    },
    {
        "key": "th17_il17",
        "label": "Th17/IL-17",
        "terms": ("Th17", "T helper 17", "IL-17", "IL-17A", "RORC", "T cell differentiation"),
        "reason": "T-cell subset bridge audit",
    },
    {
        "key": "treg_foxp3",
        "label": "Treg/FOXP3",
        "terms": ("Treg", "regulatory T cell", "FOXP3", "IL-10", "TGF-beta"),
        "reason": "T-cell subset bridge audit",
    },
    {
        "key": "macrophage_monocyte",
        "label": "macrophage/monocyte",
        "terms": (
            "macrophage",
            "monocyte",
            "macrophage polarization",
            "M1 macrophage",
            "M2 macrophage",
            "ARG1",
        ),
        "reason": "myeloid bridge audit",
    },
    {
        "key": "neutrophil_netosis",
        "label": "neutrophil/NETosis",
        "terms": ("neutrophil", "NETosis", "neutrophil extracellular trap", "CXCR2", "IL-8"),
        "reason": "neutrophil bridge audit",
    },
    {
        "key": "vascular_wall",
        "label": "endothelial/VSMC/fibroblast",
        "terms": (
            "endothelial dysfunction",
            "endothelial activation",
            "smooth muscle proliferation",
            "smooth muscle migration",
            "fibroblast activation",
            "EndMT",
        ),
        "reason": "direct vascular-wall bridge audit",
    },
    {
        "key": "ev_stromal",
        "label": "EV/stromal remodeling",
        "terms": (
            "extracellular vesicle",
            "exosome",
            "stromal remodeling",
            "extracellular matrix",
            "fibrosis",
        ),
        "reason": "vesicle or stromal bridge audit",
    },
)

STRUCTURED_BRIDGE_AXES = (
    ("mk_side", "MK/platelet/MK-EV -> candidate state", "bridge_mk_to_candidate_state"),
    (
        "metabolic_side",
        "metabolite/enzyme/product/pathway -> candidate state",
        "bridge_metabolic_to_candidate_state",
    ),
    ("disease_side", "candidate state/mediator -> PH/remodeling", "bridge_candidate_to_vascular"),
)

GENERIC_MECHANISM_TAGS = {
    "glycolysis",
    "purine metabolism",
    "pyrimidine metabolism",
}

GENERIC_HOUSEKEEPING_GENES = {
    "Gapdh",
    "Gapdhs",
    "Pkm",
    "Pgk1",
    "Tpi1",
    "Ldha",
    "Ldhb",
    "Hk1",
    "Hk2",
    "Hk3",
    "Hkdc1",
    "Gck",
    "Pfkl",
    "Pfkp",
    "Pfkm",
    "Pgam1",
    "Pgam2",
    "Pgm1",
    "Pgm2",
}

GENERIC_OR_EXOGENOUS_METABOLITES = {
    "raffinose",
    "phosphate",
}


def build_metabolic_context(
    root: Path,
    output_dir: Path,
    *,
    fetch_kegg: bool = True,
    fetch_pubmed: bool = True,
    max_metabolites: int = 30,
) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    metabolic_dir = output_dir / "metabolic"
    metabolic_dir.mkdir(parents=True, exist_ok=True)

    differentials = analyze_metabolomics(root)
    selected = select_differentials(differentials, max_metabolites=max_metabolites)
    write_differentials_csv(metabolic_dir / "differential_metabolites.csv", differentials)
    write_differentials_csv(metabolic_dir / "selected_metabolites.csv", selected)

    kegg_cache = metabolic_dir / "kegg_cache.json"
    kegg_mappings = map_metabolites_to_kegg(selected, kegg_cache, enabled=fetch_kegg)
    write_kegg_csv(metabolic_dir / "kegg_mappings.csv", kegg_mappings)

    gene_table = build_gene_table(selected, kegg_mappings)
    gene_csv = metabolic_dir / "candidate_enzyme_genes.csv"
    write_gene_csv(gene_csv, gene_table)

    expression_csv = metabolic_dir / "candidate_gene_expression.csv"
    expression_rows = score_candidate_genes(root, gene_csv, expression_csv)
    gene_table = rank_gene_table(gene_table, expression_rows)
    write_gene_csv(gene_csv, gene_table)

    pubmed_cache = metabolic_dir / "pubmed_cache.json"
    pubmed_hits = search_pubmed_for_gene_mechanisms(gene_table, pubmed_cache, enabled=fetch_pubmed)
    pubmed_csv = metabolic_dir / "pubmed_mechanism_hits.csv"
    write_pubmed_csv(pubmed_csv, pubmed_hits)

    context = render_metabolic_markdown(
        selected,
        kegg_mappings,
        gene_table,
        expression_rows,
        pubmed_hits,
        fetch_kegg=fetch_kegg,
        fetch_pubmed=fetch_pubmed,
    )
    context_path = output_dir / "metabolic_context.md"
    context_path.write_text(context, encoding="utf-8")
    manifest = {
        "generated": datetime.now().isoformat(timespec="seconds"),
        "fetch_kegg": fetch_kegg,
        "fetch_pubmed": fetch_pubmed,
        "max_metabolites": max_metabolites,
        "context_file": str(context_path),
        "supporting_files": {
            "differential_metabolites": str(metabolic_dir / "differential_metabolites.csv"),
            "selected_metabolites": str(metabolic_dir / "selected_metabolites.csv"),
            "kegg_mappings": str(metabolic_dir / "kegg_mappings.csv"),
            "candidate_enzyme_genes": str(gene_csv),
            "candidate_gene_expression": str(expression_csv),
            "pubmed_mechanism_hits": str(pubmed_csv),
        },
    }
    (metabolic_dir / "manifest.json").write_text(json.dumps(manifest, indent=2), encoding="utf-8")
    return context_path


def analyze_metabolomics(root: Path) -> list[DifferentialMetabolite]:
    results: list[DifferentialMetabolite] = []
    mk_path = root / METABOLITE_FILES["mk_metabolomics"]
    ph_path = root / METABOLITE_FILES["ph_control_metabolomics"]
    if mk_path.exists():
        results.extend(analyze_mk_workbook(mk_path))
    if ph_path.exists():
        results.extend(analyze_ph_workbook(ph_path))
    return sorted(results, key=lambda item: item.priority_score, reverse=True)


def analyze_mk_workbook(path: Path) -> list[DifferentialMetabolite]:
    import openpyxl  # type: ignore

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()
    if not rows:
        return []

    headers = [str(value) if value is not None else "" for value in rows[0]]
    groups = {
        "Control-CD41": [i for i, header in enumerate(headers) if header.startswith("Control-CD41")],
        "PH-CD41": [i for i, header in enumerate(headers) if header.startswith("PH-CD41")],
        "Control-mk": [i for i, header in enumerate(headers) if header.startswith("Control-mk")],
        "PH-mk": [i for i, header in enumerate(headers) if header.startswith("PH-mk")],
    }

    comparisons = [
        ("PH-CD41_vs_Control-CD41", groups["Control-CD41"], groups["PH-CD41"]),
        ("PH-mk_vs_Control-mk", groups["Control-mk"], groups["PH-mk"]),
    ]
    results: list[DifferentialMetabolite] = []
    for row in rows[1:]:
        metabolite = clean_name(row[0])
        if not metabolite:
            continue
        for comparison, control_idx, disease_idx in comparisons:
            control_values = numeric_values(row, control_idx)
            disease_values = numeric_values(row, disease_idx)
            if not control_values or not disease_values:
                continue
            control_mean = statistics.fmean(control_values)
            disease_mean = statistics.fmean(disease_values)
            log2fc = safe_log2_ratio(disease_mean, control_mean)
            direction = "up" if log2fc > 0 else "down"
            score = abs(log2fc) * math.log2(max(control_mean, disease_mean, 1.0) + 1.0)
            results.append(
                DifferentialMetabolite(
                    source=path.name,
                    metabolite=metabolite,
                    comparison=comparison,
                    direction=direction,
                    log2fc=log2fc,
                    control_mean=control_mean,
                    disease_mean=disease_mean,
                    priority_score=score,
                )
            )
    return results


def analyze_ph_workbook(path: Path) -> list[DifferentialMetabolite]:
    import openpyxl  # type: ignore

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    results: list[DifferentialMetabolite] = []
    for sheet_name in ("FDR", "Heatmap"):
        if sheet_name not in workbook.sheetnames:
            continue
        sheet = workbook[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(value) if value is not None else "" for value in rows[0]]
        col = {header: index for index, header in enumerate(headers)}
        required = {"WT_mean", "KO_mean", "log2FC"}
        if not required.issubset(col):
            continue
        for row in rows[1:]:
            metabolite = clean_name(row[0])
            if not metabolite:
                continue
            log2fc = safe_float(row[col["log2FC"]])
            if log2fc is None:
                continue
            wt_mean = safe_float(row[col["WT_mean"]]) or 0.0
            ko_mean = safe_float(row[col["KO_mean"]]) or 0.0
            p_value = safe_float(row[col["p_raw"]]) if "p_raw" in col else None
            fdr = safe_float(row[col["FDR_log"]]) if "FDR_log" in col else None
            fdr_bonus = 0.0 if fdr is None or fdr <= 0 else max(0.0, -math.log10(fdr))
            score = abs(log2fc) * (1.0 + fdr_bonus)
            results.append(
                DifferentialMetabolite(
                    source=f"{path.name}:{sheet_name}",
                    metabolite=metabolite,
                    comparison="KO_or_PH_vs_WT_or_control",
                    direction="up" if log2fc > 0 else "down",
                    log2fc=log2fc,
                    control_mean=wt_mean,
                    disease_mean=ko_mean,
                    p_value=p_value,
                    fdr=fdr,
                    priority_score=score,
                )
            )
    workbook.close()
    return results


def select_differentials(
    differentials: list[DifferentialMetabolite],
    *,
    max_metabolites: int,
) -> list[DifferentialMetabolite]:
    best_by_name: dict[str, DifferentialMetabolite] = {}
    for item in differentials:
        key = normalize_metabolite_key(item.metabolite)
        if key not in best_by_name or item.priority_score > best_by_name[key].priority_score:
            best_by_name[key] = item
    values = list(best_by_name.values())
    selected: list[DifferentialMetabolite] = []
    seen: set[str] = set()

    def add(items: list[DifferentialMetabolite], limit: int) -> None:
        for item in items:
            key = normalize_metabolite_key(item.metabolite)
            if key in seen:
                continue
            selected.append(item)
            seen.add(key)
            if len(selected) >= limit:
                break

    # Keep biologically interpretable KEGG-friendly small molecules from each data source.
    friendly = [item for item in values if is_kegg_friendly_metabolite(item.metabolite)]
    mk_friendly = sorted(
        [item for item in friendly if "sFig6A" in item.source],
        key=lambda item: item.priority_score,
        reverse=True,
    )
    ph_friendly = sorted(
        [item for item in friendly if "Figure6D+F" in item.source],
        key=lambda item: item.priority_score,
        reverse=True,
    )
    strong_mk_small_molecules = [
        item for item in mk_friendly
        if abs(item.log2fc) >= 2.0 and "mk_vs" in item.comparison.lower()
    ]
    add(strong_mk_small_molecules, min(max_metabolites, max(18, max_metabolites // 2)))
    add(mk_friendly, min(max_metabolites, max(20, (max_metabolites * 2) // 3)))
    add(ph_friendly, min(max_metabolites, max(12, max_metabolites // 3)))

    # Then allow high-priority complex lipids or less mappable names as secondary material.
    add(sorted(values, key=lambda item: item.priority_score, reverse=True), max_metabolites)
    return selected[:max_metabolites]


def map_metabolites_to_kegg(
    selected: list[DifferentialMetabolite],
    cache_path: Path,
    *,
    enabled: bool,
) -> dict[str, KeggMapping]:
    cache = load_json(cache_path)
    mappings: dict[str, KeggMapping] = {}
    for index, item in enumerate(selected):
        key = normalize_metabolite_key(item.metabolite)
        expand_neighbors = should_expand_pathway_neighbors(item, index)
        cached = cache.get(key)
        cache_has_neighbors = (
            isinstance(cached, dict)
            and cached.get("schema_version") == PATHWAY_NEIGHBOR_SCHEMA_VERSION
            and "neighbor_mouse_genes" in cached
            and (not expand_neighbors or bool(cached.get("neighbor_expanded")))
        )
        if (
            cached
            and cache_has_neighbors
            and not str(cached.get("status", "")).startswith("KEGG fetch failed")
        ):
            mappings[item.metabolite] = mapping_from_dict(item.metabolite, cache[key])
            continue
        if not enabled:
            mapping = KeggMapping(item.metabolite, [], [], [], [], [], [], False, "KEGG fetch disabled")
        else:
            try:
                mapping = fetch_kegg_mapping(item.metabolite, expand_neighbors=expand_neighbors)
            except Exception as exc:  # noqa: BLE001 - store a useful degradation reason.
                mapping = KeggMapping(item.metabolite, [], [], [], [], [], [], False, f"KEGG fetch failed: {exc}")
                time.sleep(0.5)
        cache[key] = mapping_to_dict(mapping)
        mappings[item.metabolite] = mapping
        save_json(cache_path, cache)
    return mappings


def should_expand_pathway_neighbors(item: DifferentialMetabolite, index: int) -> bool:
    return (
        index < 12
        and is_kegg_friendly_metabolite(item.metabolite)
        and abs(item.log2fc) >= 1.5
    )


def fetch_kegg_mapping(metabolite: str, *, expand_neighbors: bool = True) -> KeggMapping:
    find_text = kegg_get(f"/find/compound/{parse.quote(metabolite, safe='')}")
    compound_pairs = parse_kegg_pairs(find_text)
    selected_compounds = choose_compounds(metabolite, compound_pairs)
    compound_ids = [compound_id for compound_id, _ in selected_compounds[:2]]
    compound_names = [names for _, names in selected_compounds[:2]]
    if not compound_ids:
        return KeggMapping(metabolite, [], [], [], [], [], [], False, "no KEGG compound match")

    pathways: list[str] = []
    enzymes: list[str] = []
    mouse_genes: list[dict[str, str]] = []
    neighbor_mouse_genes: list[dict[str, str]] = []
    seen_genes: set[str] = set()
    for compound_id in compound_ids:
        pathways.extend(pair_right_values(kegg_get(f"/link/pathway/cpd:{compound_id}")))
        enzymes.extend(pair_right_values(kegg_get(f"/link/enzyme/cpd:{compound_id}")))
    enzymes = sorted(unique(enzymes))[:24]
    pathways = sorted(unique(pathways))[:24]

    for enzyme in enzymes[:16]:
        gene_ids = pair_right_values(kegg_get(f"/link/mmu/{enzyme}"))
        for gene_id in gene_ids[:8]:
            if gene_id in seen_genes:
                continue
            seen_genes.add(gene_id)
            gene_info = parse_kegg_gene(kegg_get(f"/get/{gene_id}"))
            gene_info["enzyme"] = enzyme
            gene_info["gene_id"] = gene_id
            gene_info["link_type"] = "direct_compound_enzyme"
            gene_info["link_pathway"] = ""
            mouse_genes.append(gene_info)
            if len(mouse_genes) >= 40:
                break
        if len(mouse_genes) >= 40:
            break

    if expand_neighbors:
        neighbor_mouse_genes = fetch_pathway_neighbor_genes(pathways, set(enzymes), seen_genes)

    status = "mapped"
    if not enzymes and not neighbor_mouse_genes:
        status = "compound mapped, no direct or pathway-neighbor enzyme links found"
    elif not mouse_genes and not neighbor_mouse_genes:
        status = "compound/enzyme mapped, no mouse genes found for linked enzymes"
    elif not enzymes and neighbor_mouse_genes:
        status = "compound mapped, pathway-neighbor enzyme genes found"
    return KeggMapping(
        metabolite,
        compound_ids,
        compound_names,
        pathways,
        enzymes,
        mouse_genes,
        neighbor_mouse_genes,
        expand_neighbors,
        status,
    )


def fetch_pathway_neighbor_genes(
    pathways: list[str],
    direct_enzymes: set[str],
    direct_gene_ids: set[str],
) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    seen_gene_ids = set(direct_gene_ids)
    for pathway in usable_neighbor_pathways(pathways)[:2]:
        mouse_pathway = mouse_pathway_id(pathway)
        pathway_name = fetch_pathway_name(pathway)
        try:
            gene_ids = pair_right_values(kegg_get(f"/link/mmu/{mouse_pathway}"))
        except Exception:
            continue
        for gene_id, gene_info in fetch_kegg_genes_batch(unique(gene_ids)[:120]):
            if gene_id in seen_gene_ids:
                continue
            enzyme = gene_info.get("enzymes", "")
            if not enzyme:
                continue
            enzyme_values = set(enzyme.split(";"))
            link_type = "pathway_neighbor_gene"
            if enzyme_values & direct_enzymes:
                link_type = "direct_pathway_gene"
            seen_gene_ids.add(gene_id)
            gene_info["enzyme"] = enzyme
            gene_info["gene_id"] = gene_id
            gene_info["link_type"] = link_type
            gene_info["link_pathway"] = pathway
            gene_info["link_pathway_name"] = pathway_name
            gene_info["mechanism_tags"] = infer_mechanism_tags(gene_info, pathway_name)
            rows.append(gene_info)
            if len(rows) >= 50:
                return rows
    return rows


def fetch_kegg_genes_batch(gene_ids: list[str], batch_size: int = 12) -> list[tuple[str, dict[str, str]]]:
    rows: list[tuple[str, dict[str, str]]] = []
    for start in range(0, len(gene_ids), batch_size):
        batch = gene_ids[start : start + batch_size]
        try:
            text = kegg_get("/get/" + "+".join(batch))
        except Exception:
            continue
        for entry in text.split("///"):
            gene_id = parse_kegg_entry_gene_id(entry)
            if not gene_id:
                continue
            rows.append((gene_id, parse_kegg_gene(entry)))
    return rows


def usable_neighbor_pathways(pathways: list[str]) -> list[str]:
    usable: list[str] = []
    for pathway in pathways:
        if pathway in GENERIC_KEGG_PATHWAYS:
            continue
        match = re.search(r"map(\d{5})$", pathway)
        if not match:
            continue
        code = int(match.group(1))
        if code <= 1099:
            usable.append(pathway)
    return unique(usable)


def mouse_pathway_id(pathway: str) -> str:
    return pathway.replace("path:map", "path:mmu", 1)


def fetch_pathway_name(pathway: str) -> str:
    try:
        return parse_kegg_entry_field(kegg_get(f"/get/{pathway}"), "NAME")
    except Exception:
        return ""


def build_gene_table(
    selected: list[DifferentialMetabolite],
    mappings: dict[str, KeggMapping],
) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    by_metabolite = {item.metabolite: item for item in selected}
    seen: set[tuple[str, str, str, str]] = set()
    for metabolite, mapping in mappings.items():
        diff = by_metabolite.get(metabolite)
        for gene in [*mapping.mouse_genes, *mapping.neighbor_mouse_genes]:
            symbol = gene.get("symbol", "")
            enzyme = gene.get("enzyme", "")
            link_type = gene.get("link_type", "direct_compound_enzyme")
            key = (normalize_metabolite_key(metabolite), enzyme, symbol.lower(), link_type)
            if not symbol or key in seen:
                continue
            seen.add(key)
            pathway_name = gene.get("link_pathway_name", "")
            mechanism_tags = gene.get("mechanism_tags", "") or infer_mechanism_tags(gene, pathway_name)
            rows.append(
                {
                    "metabolite": metabolite,
                    "source": diff.source if diff else "",
                    "comparison": diff.comparison if diff else "",
                    "metabolite_direction": diff.direction if diff else "",
                    "metabolite_log2fc": format_float(diff.log2fc) if diff else "",
                    "metabolite_priority_score": format_float(diff.priority_score) if diff else "",
                    "compound_ids": ";".join(mapping.compound_ids),
                    "pathways": ";".join(mapping.pathways[:8]),
                    "link_type": link_type,
                    "link_pathway": gene.get("link_pathway", ""),
                    "link_pathway_name": pathway_name,
                    "enzyme": enzyme,
                    "gene_id": gene.get("gene_id", ""),
                    "gene_symbol": symbol,
                    "gene_aliases": gene.get("aliases", ""),
                    "gene_name": gene.get("name", ""),
                    "gene_pathways": gene.get("pathways", ""),
                    "gene_modules": gene.get("modules", ""),
                    "mechanism_tags": mechanism_tags,
                }
            )
    return rows


def score_candidate_genes(root: Path, gene_csv: Path, output_csv: Path) -> list[dict[str, str]]:
    if not gene_csv.exists() or gene_csv.stat().st_size == 0:
        output_csv.write_text("", encoding="utf-8")
        return []
    rscript = select_rscript()
    if not rscript:
        return []
    script = root / "multi_agent_system" / "r_scripts" / "score_metabolic_genes.R"
    rds = root / "seurat_merged.rds"
    if not script.exists() or not rds.exists():
        return []
    result = subprocess.run(
        [str(rscript), str(script), str(rds), str(gene_csv), str(output_csv)],
        text=True,
        capture_output=True,
        timeout=900,
        check=False,
    )
    if result.returncode != 0:
        output_csv.write_text(
            "error\n" + (result.stdout + "\n" + result.stderr).strip(),
            encoding="utf-8",
        )
        return []
    return read_csv_dicts(output_csv)


def rank_gene_table(
    gene_table: list[dict[str, str]],
    expression_rows: list[dict[str, str]],
) -> list[dict[str, str]]:
    expression_by_gene = {row.get("gene_symbol", "").lower(): row for row in expression_rows}

    def evidence_score(row: dict[str, str]) -> float:
        value = 0.0
        log2fc = safe_float(row.get("metabolite_log2fc"))
        priority = safe_float(row.get("metabolite_priority_score"))
        if log2fc is not None:
            value += min(abs(log2fc), 8.0) * 2.2
        if priority is not None:
            value += min(priority / 8.0, 8.0)

        link_type = row.get("link_type", "")
        if link_type == "direct_compound_enzyme":
            value += 2.2
        elif link_type == "direct_pathway_gene":
            value += 2.0
        elif link_type == "pathway_neighbor_gene":
            value += 1.6

        if row.get("mechanism_tags"):
            value += 0.8
        if row.get("gene_modules"):
            value += 0.6

        expr = expression_by_gene.get(row.get("gene_symbol", "").lower(), {})
        if expr.get("status") == "matched":
            value += 3.0
            mk_pct = safe_float(expr.get("mk_pct_expr"))
            mk_enrichment = safe_float(expr.get("mk_enrichment_log2"))
            ph_log2 = safe_float(expr.get("ph_vs_control_mk_log2"))
            p_value = safe_float(expr.get("ph_vs_control_mk_p_value"))
            if mk_pct is not None:
                value += min(mk_pct / 10.0, 4.0)
            if mk_enrichment is not None and mk_enrichment > 0:
                value += min(mk_enrichment, 3.0)
            if ph_log2 is not None:
                value += min(abs(ph_log2), 3.0)
                if ph_log2 > 0:
                    value += 0.5
            if p_value is not None and p_value < 0.05:
                value += 2.0
        return value

    def readiness_score(row: dict[str, str]) -> float:
        value = 0.0
        log2fc = safe_float(row.get("metabolite_log2fc"))
        if log2fc is not None:
            value += min(abs(log2fc), 4.0) * 0.8

        if "sFig6A" in row.get("source", "") and "mk_vs" in row.get("comparison", "").lower():
            value += 2.0

        link_type = row.get("link_type", "")
        if link_type == "direct_compound_enzyme":
            value += 1.8
        elif link_type == "direct_pathway_gene":
            value += 1.6
        elif link_type == "pathway_neighbor_gene":
            value += 1.2

        tags = mechanism_tag_set(row)
        strong_tags = tags & STRONG_MECHANISM_TAGS
        generic_tags = tags & GENERIC_MECHANISM_TAGS
        if strong_tags:
            value += min(len(strong_tags) * 3.0, 10.0)
        if row.get("gene_modules"):
            value += 2.0
        if not tags:
            value -= 1.0
        elif generic_tags and not strong_tags:
            value -= 4.0

        symbol = row.get("gene_symbol", "")
        if symbol in GENERIC_HOUSEKEEPING_GENES and not strong_tags:
            value -= 4.0

        metabolite_key = normalize_metabolite_key(row.get("metabolite", ""))
        if metabolite_key in GENERIC_OR_EXOGENOUS_METABOLITES and not strong_tags:
            value -= 5.0

        expr = expression_by_gene.get(symbol.lower(), {})
        if expr.get("status") != "matched":
            return value - 4.0

        value += 1.0
        mk_pct = safe_float(expr.get("mk_pct_expr"))
        mk_enrichment = safe_float(expr.get("mk_enrichment_log2"))
        ph_log2 = safe_float(expr.get("ph_vs_control_mk_log2"))
        p_value = safe_float(expr.get("ph_vs_control_mk_p_value"))

        if mk_pct is not None:
            if mk_pct >= 20:
                value += 3.0
            elif mk_pct >= 10:
                value += 2.0
            elif mk_pct >= 5:
                value += 1.0
            elif mk_pct < 2:
                value -= 1.5

        if mk_enrichment is not None:
            if mk_enrichment > 0:
                value += min(mk_enrichment * 3.0, 6.0)
                if mk_enrichment >= 1:
                    value += 1.0
            elif mk_enrichment < -1:
                value -= 4.0

        if ph_log2 is not None:
            if ph_log2 > 0:
                value += min(ph_log2 * 2.0, 5.0)
                if ph_log2 >= 1:
                    value += 1.0
            else:
                value -= 2.0

        if p_value is not None:
            if p_value < 0.01:
                value += 4.0
            elif p_value < 0.05:
                value += 3.0
            elif p_value > 0.2:
                value -= 1.0

        return value

    ranked = sorted(gene_table, key=readiness_score, reverse=True)
    for row in ranked:
        row["evidence_priority_score"] = format_float(evidence_score(row))
        row["hypothesis_readiness_score"] = format_float(readiness_score(row))
    return ranked


def mechanism_tag_set(row: dict[str, str]) -> set[str]:
    return {
        tag.strip().lower()
        for tag in row.get("mechanism_tags", "").split(";")
        if tag.strip()
    }


def search_pubmed_for_gene_mechanisms(
    gene_table: list[dict[str, str]],
    cache_path: Path,
    *,
    enabled: bool,
) -> list[dict[str, str]]:
    cache = load_json(cache_path)
    hits: list[dict[str, str]] = []
    gene_rows: dict[str, list[dict[str, str]]] = {}
    seen: set[str] = set()
    for row in gene_table:
        symbol = row.get("gene_symbol", "")
        if symbol and symbol.lower() not in seen:
            gene_rows[symbol] = []
            seen.add(symbol.lower())
        if symbol in gene_rows:
            gene_rows[symbol].append(row)
        if len(gene_rows) >= MAX_PUBMED_GENES:
            break
    for symbol, rows_for_gene in gene_rows.items():
        context_terms = pubmed_context_terms(rows_for_gene)
        bridge_terms = pubmed_bridge_terms(rows_for_gene)
        cache_key = "v8|" + symbol + "|" + ",".join(context_terms[:4]) + "|" + ",".join(bridge_terms[:6])
        if cache_key in cache:
            hits.extend(cache[cache_key])
            continue
        if not enabled:
            cache[cache_key] = []
            continue
        try:
            rows = fetch_pubmed_hits(symbol, context_terms)
            rows.extend(fetch_context_pubmed_hits(symbol, context_terms))
            rows.extend(fetch_bridge_candidate_pubmed_hits(symbol, bridge_terms))
            rows.extend(fetch_bridge_to_vascular_hits(symbol, bridge_terms))
            rows.extend(fetch_structured_bridge_convergence_hits(symbol, rows_for_gene, bridge_terms))
        except Exception:
            rows = []
        cache[cache_key] = rows
        hits.extend(rows)
        save_json(cache_path, cache)
        time.sleep(0.35)
    return hits


def pubmed_title_abstract_or(terms: list[str] | tuple[str, ...]) -> str:
    parts: list[str] = []
    for term in unique([item for item in terms if item]):
        cleaned = term.replace('"', "").strip()
        if not cleaned:
            continue
        if " " in cleaned or "-" in cleaned:
            parts.append(f'"{cleaned}"[Title/Abstract]')
        else:
            parts.append(f"{cleaned}[Title/Abstract]")
    return " OR ".join(parts)


def pubmed_search_url(term: str, retmax: int = 3) -> str:
    return (
        f"{NCBI_BASE}/esearch.fcgi?db=pubmed&retmode=json&sort=relevance&retmax={retmax}&term="
        + parse.quote(term)
    )


def pubmed_summary_rows(
    *,
    symbol: str,
    ids: list[str],
    query_terms: list[str],
    hit_type: str,
    bridge_candidate: str = "",
    evidence_role: str = "",
) -> list[dict[str, str]]:
    if not ids:
        return []
    summary_url = (
        f"{NCBI_BASE}/esummary.fcgi?db=pubmed&retmode=json&id="
        + ",".join(ids)
    )
    summary_data = json.loads(url_get(summary_url))
    rows: list[dict[str, str]] = []
    for pmid in ids:
        item = summary_data.get("result", {}).get(pmid, {})
        title = item.get("title", "").strip()
        if not title:
            continue
        rows.append(
            {
                "gene_symbol": symbol,
                "pmid": pmid,
                "title": title,
                "journal": item.get("fulljournalname", item.get("source", "")),
                "pubdate": item.get("pubdate", ""),
                "query_terms": ";".join(unique(query_terms)),
                "hit_type": hit_type,
                "bridge_candidate": bridge_candidate,
                "evidence_role": evidence_role,
            }
        )
    return rows


def fetch_pubmed_query_rows(
    *,
    symbol: str,
    term: str,
    query_terms: list[str],
    hit_type: str,
    bridge_candidate: str = "",
    evidence_role: str = "",
    retmax: int = 3,
) -> list[dict[str, str]]:
    search_data = json.loads(url_get(pubmed_search_url(term, retmax=retmax)))
    ids = search_data.get("esearchresult", {}).get("idlist", [])
    rows = pubmed_summary_rows(
        symbol=symbol,
        ids=ids,
        query_terms=query_terms,
        hit_type=hit_type,
        bridge_candidate=bridge_candidate,
        evidence_role=evidence_role,
    )
    return rank_pubmed_rows(rows, query_terms, evidence_role)


def rank_pubmed_rows(rows: list[dict[str, str]], query_terms: list[str], evidence_role: str) -> list[dict[str, str]]:
    def score(row: dict[str, str]) -> tuple[int, str]:
        title = row.get("title", "").lower()
        value = 0
        for term in query_terms:
            normalized = term.lower().replace('"', "").strip()
            if normalized and normalized in title:
                value += 4 if " " in normalized else 2
        if "pulmonary arterial hypertension" in title or "pulmonary hypertension" in title:
            value += 10
        if "pulmonary vascular" in title:
            value += 6
        if (
            "polyamine" in title
            or "spermidine" in title
            or "spermine" in title
            or "methionine" in title
            or "arginine" in title
            or "ornithine" in title
            or "tryptophan" in title
            or "kynurenine" in title
            or "retinoic" in title
            or "glutathione" in title
            or "methylation" in title
        ):
            value += 6
        if "megakaryocyte" in title or "platelet-derived" in title or "platelet extracellular" in title:
            value += 6
        if evidence_role.startswith("candidate state") and "pulmonary" not in title:
            value -= 5
        if evidence_role.startswith("MK/platelet") and not (
            "megakaryocyte" in title or "platelet" in title
        ):
            value -= 3
        return value, row.get("pubdate", "")

    return sorted(rows, key=score, reverse=True)


def fetch_structured_bridge_convergence_hits(
    symbol: str,
    rows_for_gene: list[dict[str, str]],
    bridge_terms: list[str],
) -> list[dict[str, str]]:
    if not bridge_terms:
        return []
    if not should_fetch_structured_bridge(rows_for_gene):
        return []
    hits: list[dict[str, str]] = []
    metabolic_terms = unique([symbol] + bridge_terms[:8])
    mk_query = pubmed_title_abstract_or(MK_BRIDGE_SOURCE_TERMS)
    vascular_query = pubmed_title_abstract_or(STRUCTURED_DISEASE_BRIDGE_TERMS)
    metabolic_query = pubmed_title_abstract_or(metabolic_terms)
    if not mk_query or not vascular_query or not metabolic_query:
        return []
    for profile in STRUCTURED_BRIDGE_PROFILES:
        label = str(profile["label"])
        candidate_terms = list(profile["terms"])
        candidate_query = pubmed_title_abstract_or(tuple(candidate_terms))
        if not candidate_query:
            continue
        axis_terms = {
            "mk_side": f"({mk_query}) AND ({candidate_query})",
            "metabolic_side": f"({metabolic_query}) AND ({candidate_query})",
            "disease_side": f"({candidate_query}) AND ({vascular_query})",
        }
        query_term_map = {
            "mk_side": list(MK_BRIDGE_SOURCE_TERMS) + candidate_terms,
            "metabolic_side": metabolic_terms + candidate_terms,
            "disease_side": candidate_terms + list(STRUCTURED_DISEASE_BRIDGE_TERMS),
        }
        for axis_key, evidence_role, hit_type in STRUCTURED_BRIDGE_AXES:
            hits.extend(
                fetch_pubmed_query_rows(
                    symbol=symbol,
                    term=axis_terms[axis_key],
                    query_terms=query_term_map[axis_key],
                    hit_type=hit_type,
                    bridge_candidate=label,
                    evidence_role=evidence_role,
                    retmax=STRUCTURED_BRIDGE_RETRIEVE_PER_AXIS,
                )
            )
            time.sleep(0.15)
    return dedupe_pubmed_hits(hits)


def should_fetch_structured_bridge(rows_for_gene: list[dict[str, str]]) -> bool:
    best_readiness = max(
        (safe_float(row.get("hypothesis_readiness_score")) or 0.0 for row in rows_for_gene),
        default=0.0,
    )
    tags = set()
    for row in rows_for_gene:
        tags |= mechanism_tag_set(row)
    return best_readiness >= STRUCTURED_BRIDGE_READINESS_MIN and bool(tags & STRONG_MECHANISM_TAGS)


def dedupe_pubmed_hits(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    deduped: list[dict[str, str]] = []
    seen: set[tuple[str, str, str, str]] = set()
    for row in rows:
        key = (
            row.get("pmid", ""),
            row.get("hit_type", ""),
            row.get("bridge_candidate", ""),
            row.get("evidence_role", ""),
        )
        if not key[0] or key in seen:
            continue
        seen.add(key)
        deduped.append(row)
    return deduped


def fetch_pubmed_hits(symbol: str, context_terms: list[str]) -> list[dict[str, str]]:
    remodeling_terms = (
        '"vascular remodeling"[Title/Abstract] OR "pulmonary hypertension"[Title/Abstract] OR '
        'hypoxia[Title/Abstract] OR endothelial[Title/Abstract] OR "smooth muscle"[Title/Abstract] OR '
        'fibroblast[Title/Abstract] OR inflammation[Title/Abstract]'
    )
    mechanism_terms = " OR ".join(
        f'"{term}"[Title/Abstract]' if " " in term else f"{term}[Title/Abstract]"
        for term in context_terms[:4]
    )
    evidence_terms = remodeling_terms
    if mechanism_terms:
        evidence_terms = f"({remodeling_terms} OR {mechanism_terms})"
    term = (
        f'("{symbol}"[Title/Abstract]) AND '
        f"{evidence_terms}"
    )
    search_url = (
        pubmed_search_url(term, retmax=4)
    )
    search_data = json.loads(url_get(search_url))
    ids = search_data.get("esearchresult", {}).get("idlist", [])
    return pubmed_summary_rows(
        symbol=symbol,
        ids=ids,
        query_terms=context_terms[:4],
        hit_type="gene_or_gene_context",
    )


def fetch_bridge_candidate_pubmed_hits(symbol: str, bridge_terms: list[str]) -> list[dict[str, str]]:
    if not bridge_terms:
        return []
    metabolic_query = pubmed_title_abstract_or([symbol] + bridge_terms[:6])
    candidate_query = pubmed_title_abstract_or(CANDIDATE_BRIDGE_TERMS)
    if not metabolic_query or not candidate_query:
        return []
    term = f"({metabolic_query}) AND ({candidate_query})"
    search_url = (
        pubmed_search_url(term, retmax=3)
    )
    search_data = json.loads(url_get(search_url))
    ids = search_data.get("esearchresult", {}).get("idlist", [])
    return pubmed_summary_rows(
        symbol=symbol,
        ids=ids,
        query_terms=bridge_terms[:6] + ["balanced bridge candidate screen"],
        hit_type="bridge_metabolite_to_candidate_state",
    )


def fetch_bridge_to_vascular_hits(symbol: str, bridge_terms: list[str]) -> list[dict[str, str]]:
    if not bridge_terms:
        return []
    candidate_query = pubmed_title_abstract_or(CANDIDATE_BRIDGE_TERMS)
    vascular_query = pubmed_title_abstract_or(VASCULAR_BRIDGE_TERMS)
    if not candidate_query or not vascular_query:
        return []
    term = f"({candidate_query}) AND ({vascular_query})"
    search_url = (
        pubmed_search_url(term, retmax=2)
    )
    search_data = json.loads(url_get(search_url))
    ids = search_data.get("esearchresult", {}).get("idlist", [])
    return pubmed_summary_rows(
        symbol=symbol,
        ids=ids,
        query_terms=["balanced bridge candidate screen", "pulmonary hypertension", "vascular remodeling"],
        hit_type="bridge_candidate_state_to_vascular",
    )


def fetch_context_pubmed_hits(symbol: str, context_terms: list[str]) -> list[dict[str, str]]:
    if not context_terms:
        return []
    remodeling_terms = (
        '"vascular remodeling"[Title/Abstract] OR "pulmonary hypertension"[Title/Abstract] OR '
        'hypoxia[Title/Abstract] OR endothelial[Title/Abstract] OR "smooth muscle"[Title/Abstract] OR '
        'fibroblast[Title/Abstract] OR inflammation[Title/Abstract]'
    )
    mechanism_terms = " OR ".join(
        f'"{term}"[Title/Abstract]' if " " in term else f"{term}[Title/Abstract]"
        for term in context_terms[:4]
    )
    term = f"({mechanism_terms}) AND ({remodeling_terms})"
    search_url = (
        pubmed_search_url(term, retmax=2)
    )
    search_data = json.loads(url_get(search_url))
    ids = search_data.get("esearchresult", {}).get("idlist", [])
    return pubmed_summary_rows(
        symbol=symbol,
        ids=ids,
        query_terms=context_terms[:4],
        hit_type="mechanism_context",
    )


def pubmed_context_terms(rows: list[dict[str, str]]) -> list[str]:
    terms: list[str] = []
    text_parts: list[str] = []
    for row in rows:
        text_parts.extend(
            [
                row.get("mechanism_tags", ""),
                row.get("gene_name", ""),
                row.get("gene_aliases", ""),
                row.get("gene_pathways", ""),
                row.get("gene_modules", ""),
                row.get("link_pathway_name", ""),
            ]
        )
    text = " ".join(text_parts).lower()
    for term in (
        "pulmonary hypertension",
        "vascular remodeling",
        "hypoxia",
        "endothelial",
        "smooth muscle",
        "polyamine",
        "methionine salvage",
        "s-adenosylmethionine",
        "glutathione",
        "one-carbon",
        "tryptophan",
        "retinoic acid",
    ):
        if term in text:
            terms.append(term)
    return unique(terms)


def pubmed_bridge_terms(rows: list[dict[str, str]]) -> list[str]:
    terms: list[str] = []
    text_parts: list[str] = []
    for row in rows:
        text_parts.extend(
            [
                row.get("metabolite", ""),
                row.get("mechanism_tags", ""),
                row.get("gene_name", ""),
                row.get("gene_aliases", ""),
                row.get("gene_pathways", ""),
                row.get("gene_modules", ""),
                row.get("link_pathway_name", ""),
            ]
        )
    text = " ".join(text_parts).lower()
    if "methionine" in text:
        terms.extend(["methionine", "S-adenosylmethionine", "methionine salvage"])
    if "adenosylmethionine" in text or "sam" in text:
        terms.extend(["S-adenosylmethionine"])
    if "polyamine" in text or "decarboxylase" in text:
        terms.extend(["polyamine", "spermidine", "spermine"])
    if "one-carbon" in text or "one carbon" in text:
        terms.extend(["one-carbon metabolism", "methylation"])
    if "arginine" in text or "proline" in text:
        terms.extend(["arginine", "ornithine", "polyamine"])
    if "tryptophan" in text:
        terms.extend(["tryptophan", "kynurenine"])
    if "retinoic" in text or "retinoid" in text:
        terms.extend(["retinoic acid", "retinoid"])
    if "glutathione" in text or "redox" in text:
        terms.extend(["glutathione", "redox"])
    return unique(terms)


def format_pubmed_hits(hits: list[dict[str, str]], limit: int = 3) -> str:
    if not hits:
        return ""
    hit_priority = {
        "gene_or_gene_context": 0,
        "mechanism_context": 1,
        "bridge_metabolic_to_candidate_state": 2,
        "bridge_candidate_state_to_vascular": 3,
        "bridge_mk_to_candidate_state": 4,
        "bridge_candidate_to_vascular": 5,
        "bridge_metabolite_to_candidate_state": 6,
    }
    ordered = sorted(
        enumerate(hits),
        key=lambda item: (hit_priority.get(item[1].get("hit_type", ""), 9), item[0]),
    )
    selected: list[dict[str, str]] = []
    used_pmids: set[str] = set()
    for _, hit in ordered:
        if len(selected) >= limit:
            break
        if hit.get("pmid") in used_pmids:
            continue
        selected.append(hit)
        used_pmids.add(hit.get("pmid", ""))
    summaries: list[str] = []
    for hit in selected[:limit]:
        hit_type = hit.get("hit_type", "hit")
        candidate = hit.get("bridge_candidate", "")
        label = f"{hit_type}({candidate})" if candidate else hit_type
        summaries.append(f"{label} PMID {hit['pmid']}: {hit['title'][:90]}")
    return "; ".join(summaries)


def format_bridge_axis_hits(
    hits: list[dict[str, str]],
    candidate: str,
    evidence_role: str,
    limit: int = 2,
) -> str:
    selected = select_bridge_axis_hits(hits, candidate, evidence_role, limit=limit)
    if not selected:
        return "none retrieved"
    return "; ".join(
        f"PMID {hit.get('pmid', '')}: {hit.get('title', '')[:110]}"
        for hit in selected
    ) or "none retrieved"


def select_bridge_axis_hits(
    hits: list[dict[str, str]],
    candidate: str,
    evidence_role: str,
    limit: int | None = None,
) -> list[dict[str, str]]:
    selected = [
        hit
        for hit in hits
        if hit.get("bridge_candidate") == candidate and hit.get("evidence_role") == evidence_role
    ]
    deduped: list[dict[str, str]] = []
    used_pmids: set[str] = set()
    for hit in selected:
        pmid = hit.get("pmid", "")
        if not pmid or pmid in used_pmids:
            continue
        used_pmids.add(pmid)
        deduped.append(hit)
        if limit is not None and len(deduped) >= limit:
            break
    return deduped


def bridge_candidate_support_summary(hits: list[dict[str, str]], candidate: str) -> tuple[int, int, str]:
    supported_roles: list[str] = []
    total_hits = 0
    for _axis_key, evidence_role, _hit_type in STRUCTURED_BRIDGE_AXES:
        axis_hits = select_bridge_axis_hits(hits, candidate, evidence_role)
        if axis_hits:
            supported_roles.append(evidence_role)
            total_hits += len(axis_hits)
    return len(supported_roles), total_hits, "; ".join(supported_roles)


def bridge_candidates_for_display(
    hits: list[dict[str, str]],
    limit: int = STRUCTURED_BRIDGE_DISPLAY_CANDIDATES_PER_CHAIN,
) -> list[str]:
    scored: list[tuple[int, int, int, str]] = []
    for index, profile in enumerate(STRUCTURED_BRIDGE_PROFILES):
        candidate = str(profile["label"])
        supported_axes, total_hits, _roles = bridge_candidate_support_summary(hits, candidate)
        if supported_axes == 0:
            continue
        scored.append((supported_axes, total_hits, -index, candidate))
    scored.sort(reverse=True)
    return [candidate for _axes, _hits, _index, candidate in scored[:limit]]


def render_metabolic_markdown(
    selected: list[DifferentialMetabolite],
    mappings: dict[str, KeggMapping],
    gene_table: list[dict[str, str]],
    expression_rows: list[dict[str, str]],
    pubmed_hits: list[dict[str, str]],
    *,
    fetch_kegg: bool,
    fetch_pubmed: bool,
) -> str:
    expression_by_gene = {row.get("gene_symbol", "").lower(): row for row in expression_rows}
    pubmed_by_gene: dict[str, list[dict[str, str]]] = {}
    for hit in pubmed_hits:
        pubmed_by_gene.setdefault(hit.get("gene_symbol", "").lower(), []).append(hit)
    ready_rows = [
        row for row in gene_table
        if is_mechanism_ready(row, expression_by_gene.get(row.get("gene_symbol", "").lower(), {}))
    ][:15]

    lines = [
        "# Metabolomics-to-Mechanism Evidence Context",
        "",
        "This context is generated before hypothesis generation to support metabolomics-driven, direction-level mechanisms.",
        "Evidence chain target: differential metabolite -> KEGG compound/pathway/enzyme or pathway-neighbor gene -> candidate enzyme gene -> MK expression/differential evidence -> directional downstream biology -> vascular remodeling phenotype.",
        "",
        f"- KEGG fetch: {'enabled' if fetch_kegg else 'disabled'}",
        f"- PubMed fetch: {'enabled' if fetch_pubmed else 'disabled'}",
        f"- Selected metabolites: {len(selected)}",
        f"- Candidate enzyme-gene rows: {len(gene_table)}",
        f"- Direct compound-enzyme rows: {sum(1 for row in gene_table if row.get('link_type') == 'direct_compound_enzyme')}",
        f"- Pathway-neighbor rows: {sum(1 for row in gene_table if row.get('link_type') == 'pathway_neighbor_gene')}",
        f"- Candidate genes with Seurat expression metrics: {len(expression_rows)}",
        f"- Mechanism-ready shortlist rows: {len(ready_rows)}",
        "",
        "## Mechanism-Ready Hypothesis Shortlist",
        "These rows are ranked for hypothesis generation, not just KEGG coverage. They require a differential MK metabolite, a KEGG direct or same-pathway enzyme/gene link, matched Seurat expression, positive MK enrichment, PH-up shift in MKs, and non-generic mechanism cues. Generation agents should use top rows as direction-level anchors, not as proof of a specific downstream cell subtype, cytokine, or final bridge.",
        "",
        "| Rank | Metabolite | Candidate gene | Readiness | Link type | KEGG/mechanism context | Seurat MK/PH evidence | Literature hits |",
        "|---:|---|---|---:|---|---|---|---|",
    ]
    for index, row in enumerate(ready_rows, start=1):
        symbol = row.get("gene_symbol", "")
        expr = expression_by_gene.get(symbol.lower(), {})
        pubmed = pubmed_by_gene.get(symbol.lower(), [])
        pubmed_text = format_pubmed_hits(pubmed, limit=3) or "none retrieved"
        kegg_context = "; ".join(
            value
            for value in [
                row.get("enzyme", ""),
                row.get("link_pathway_name", ""),
                row.get("gene_modules", ""),
                row.get("mechanism_tags", ""),
            ]
            if value
        )
        lines.append(
            "| "
            + " | ".join(
                [
                    str(index),
                    md_escape(
                        f"{row.get('metabolite', '')} "
                        f"({row.get('metabolite_direction', '')}, log2FC {row.get('metabolite_log2fc', '')})"
                    ),
                    md_escape(f"{symbol}: {row.get('gene_name', '')}"),
                    md_escape(row.get("hypothesis_readiness_score", "")),
                    md_escape(row.get("link_type", "")),
                    md_escape(kegg_context),
                    md_escape(expression_summary(expr)),
                    md_escape(pubmed_text),
                ]
            )
            + " |"
        )

    bridge_rows = [
        (row, pubmed_bridge_terms([row]))
        for row in ready_rows
        if pubmed_bridge_terms([row])
    ]
    if bridge_rows:
        lines.extend(
            [
                "",
                "## Directional Downstream Axis Cues",
                "For these mechanism-ready chains, generation agents should name plausible downstream axes without selecting a final bridge unless direct evidence supports that specificity. Treat immune subsets, cytokines, vascular recipients, EV routes, and stromal routes as candidate examples rather than settled mechanisms.",
                "Use these cues to keep hypotheses biologically oriented: MK metabolic state -> pathway class -> broad downstream axis -> vascular remodeling phenotype. Avoid over-resolving to a single metabolite product, T-cell subset, mediator, or EndMT route.",
            ]
        )
        for row, terms in bridge_rows[:8]:
            symbol = row.get("gene_symbol", "")
            axis_focus = (
                "candidate downstream axes: immune-mediated or T-helper/Th17-like tone; "
                "macrophage/monocyte or neutrophil inflammation; direct endothelial/smooth-muscle/"
                "fibroblast activation; EV/stromal remodeling; unresolved"
            )
            lines.append(
                "- "
                + md_escape(
                    f"{row.get('metabolite', '')} -> {symbol}: pathway terms {', '.join(terms[:8])}; "
                    f"{axis_focus}"
                )
            )

    lines.extend(
        [
            "",
        "## Prioritized Complete Evidence Chains",
        "| Metabolite | Metabolite signal | Link type | Candidate gene | KEGG context | Seurat MK/PH evidence | Mechanism cues | Literature hits |",
        "|---|---|---|---|---|---|---|---|",
        ]
    )
    for row in gene_table[:25]:
        symbol = row.get("gene_symbol", "")
        expr = expression_by_gene.get(symbol.lower(), {})
        pubmed = pubmed_by_gene.get(symbol.lower(), [])
        pubmed_text = format_pubmed_hits(pubmed, limit=3) or "none retrieved"
        kegg_context = "; ".join(
            value
            for value in [
                row.get("enzyme", ""),
                row.get("link_pathway_name", ""),
                row.get("gene_modules", ""),
            ]
            if value
        )
        lines.append(
            "| "
            + " | ".join(
                [
                    md_escape(row.get("metabolite", "")),
                    md_escape(
                        f"{row.get('metabolite_direction', '')} log2FC {row.get('metabolite_log2fc', '')} "
                        f"({row.get('comparison', '')})"
                    ),
                    md_escape(row.get("link_type", "")),
                    md_escape(
                        f"{symbol}: {row.get('gene_name', '')} "
                        f"[readiness {row.get('hypothesis_readiness_score', '')}; evidence {row.get('evidence_priority_score', '')}]"
                    ),
                    md_escape(kegg_context or row.get("pathways", "")),
                    md_escape(expression_summary(expr)),
                    md_escape(row.get("mechanism_tags", "") or "none inferred"),
                    md_escape(pubmed_text),
                ]
            )
            + " |"
        )

    lines.extend(
        [
            "",
        "## Priority Differential Metabolites",
        "| Metabolite | Source | Comparison | Direction | log2FC | FDR | Priority |",
        "|---|---|---|---|---:|---:|---:|",
        ]
    )
    for item in selected[:30]:
        lines.append(
            "| "
            + " | ".join(
                [
                    md_escape(item.metabolite),
                    md_escape(item.source),
                    md_escape(item.comparison),
                    item.direction,
                    format_float(item.log2fc),
                    format_float(item.fdr),
                    format_float(item.priority_score),
                ]
            )
            + " |"
        )

    lines.extend(["", "## KEGG Metabolite-Enzyme Mapping"])
    for item in selected[:20]:
        mapping = mappings.get(item.metabolite)
        if not mapping:
            continue
        lines.extend(
            [
                "",
                f"### {item.metabolite}",
                f"- Differential signal: {item.direction}, log2FC {format_float(item.log2fc)} in {item.comparison} ({item.source})",
                f"- KEGG status: {mapping.status}",
                f"- KEGG compounds: {', '.join(mapping.compound_ids) or 'none'}",
                f"- KEGG compound names: {', '.join(mapping.compound_names[:3]) or 'none'}",
                f"- KEGG pathways: {', '.join(mapping.pathways[:10]) or 'none'}",
                f"- Linked enzymes: {', '.join(mapping.enzymes[:12]) or 'none'}",
            ]
        )
        rows = [row for row in gene_table if row["metabolite"] == item.metabolite]
        if not rows:
            lines.append("- Candidate mouse enzyme genes: none mapped")
            continue
        lines.append("- Candidate mouse enzyme genes with MK expression evidence:")
        for row in rows[:10]:
            symbol = row["gene_symbol"]
            expr = expression_by_gene.get(symbol.lower(), {})
            expr_text = expression_summary(expr)
            link_context = "; ".join(
                value
                for value in [
                    row.get("link_type", ""),
                    row.get("link_pathway_name", ""),
                    row.get("gene_modules", ""),
                    row.get("mechanism_tags", ""),
                ]
                if value
            )
            pubmed = pubmed_by_gene.get(symbol.lower(), [])
            pubmed_text = format_pubmed_hits(pubmed, limit=3)
            lines.append(
                f"  - {symbol} ({row['enzyme']}): {row.get('gene_name','')}; {link_context}; {expr_text}"
                + (f"; PubMed hits: {pubmed_text}" if pubmed_text else "; PubMed hits: none retrieved")
            )

    lines.extend(
        [
            "",
            "## Required Hypothesis Scaffold For Generation Agents",
            "Every metabolomics-driven hypothesis must include:",
            "- Differential metabolite A with source/comparison/log2FC.",
            "- KEGG direct compound-enzyme evidence or a same-pathway neighbor-gene link for A.",
            "- Candidate metabolic enzyme gene B and whether B is expressed/enriched/differential in MK/platelet cells.",
            "- The evidence link type, especially whether B is a direct compound-enzyme link or a pathway-neighbor gene.",
            "- Whether literature search hits support a vascular remodeling, hypoxia, endothelial, smooth muscle, fibroblast, immune, or pulmonary hypertension mechanism.",
            "- A direction-level chain from hypoxic MK metabolic state to broad downstream axis to remodeling phenotype.",
            "- A direction-level reasoning summary that links the data anchor, biological interpretation, MK-linked enzyme/pathway logic, plausible downstream axis, remodeling phenotype, and key uncertainty.",
            "- A broad downstream axis: direct vascular-wall, immune-mediated, EV/stromal, or unresolved.",
            "- Candidate examples can include SAM, spermidine, T-helper/Th17-like tone, endothelial activation, medial activation, muscularization, vascular stiffness, or stromal remodeling, but examples should not be presented as settled mechanisms.",
            "- A falsification test that can challenge the metabolite-enzyme-MK-remodeling direction.",
            "",
            "Do not present KEGG mapping or PubMed hits as proof of causality. Treat them as evidence scaffolds requiring validation, and avoid over-resolving the final mechanism when evidence only supports a direction.",
        ]
    )
    return "\n".join(lines).strip() + "\n"


def expression_summary(row: dict[str, str]) -> str:
    if not row:
        return "Seurat expression: not checked"
    if row.get("status") != "matched":
        return f"Seurat expression: {row.get('status', 'not matched')}"
    return (
        f"Seurat expression: MK pct {row.get('mk_pct_expr','NA')}%, "
        f"other pct {row.get('other_pct_expr','NA')}%, "
        f"MK enrichment log2 {row.get('mk_enrichment_log2','NA')}, "
        f"PH-vs-control MK log2 {row.get('ph_vs_control_mk_log2','NA')}, "
        f"PH MK pct {row.get('ph_mk_pct_expr','NA')}%, "
        f"control MK pct {row.get('control_mk_pct_expr','NA')}%, "
        f"PH-vs-control MK Wilcoxon p {row.get('ph_vs_control_mk_p_value','NA')}"
    )


def is_mechanism_ready(row: dict[str, str], expression: dict[str, str]) -> bool:
    if expression.get("status") != "matched":
        return False
    mk_enrichment = safe_float(expression.get("mk_enrichment_log2"))
    ph_log2 = safe_float(expression.get("ph_vs_control_mk_log2"))
    mk_pct = safe_float(expression.get("mk_pct_expr"))
    if mk_enrichment is None or mk_enrichment <= 0:
        return False
    if ph_log2 is None or ph_log2 <= 0.5:
        return False
    if mk_pct is None or mk_pct < 2:
        return False
    if not (mechanism_tag_set(row) & STRONG_MECHANISM_TAGS):
        return False
    if row.get("gene_symbol", "") in GENERIC_HOUSEKEEPING_GENES:
        return False
    return True


def kegg_get(path: str) -> str:
    return url_get(KEGG_BASE + path)


def url_get(url: str) -> str:
    with request.urlopen(url, timeout=30) as response:
        return response.read().decode("utf-8", errors="replace")


def parse_kegg_pairs(text: str) -> list[tuple[str, str]]:
    pairs: list[tuple[str, str]] = []
    for line in text.splitlines():
        if "\t" not in line:
            continue
        left, right = line.split("\t", 1)
        pairs.append((left.replace("cpd:", ""), right.strip()))
    return pairs


def pair_right_values(text: str) -> list[str]:
    values: list[str] = []
    for line in text.splitlines():
        if "\t" not in line:
            continue
        _, right = line.split("\t", 1)
        values.append(right.strip())
    return values


def choose_compounds(metabolite: str, pairs: list[tuple[str, str]]) -> list[tuple[str, str]]:
    key = normalize_metabolite_key(metabolite)

    def score(pair: tuple[str, str]) -> tuple[int, int]:
        names = [normalize_metabolite_key(name) for name in re.split(r";\s*", pair[1])]
        exact = 1 if key in names else 0
        starts = 1 if any(name.startswith(key) or key.startswith(name) for name in names) else 0
        return exact + starts, -len(pair[1])

    return sorted(pairs, key=score, reverse=True)


def parse_kegg_gene(text: str) -> dict[str, str]:
    symbol = ""
    aliases = ""
    names: list[str] = []
    pathways: list[str] = []
    modules: list[str] = []
    orthology: list[str] = []
    enzymes: list[str] = []
    current_field = ""
    for line in text.splitlines():
        field = line[:12].strip()
        value = line[12:].strip() if len(line) > 12 else ""
        if field:
            current_field = field
        else:
            field = current_field
        if field == "SYMBOL":
            parts = [part.strip() for part in value.split(",") if part.strip()]
            symbol = parts[0] if parts else ""
            aliases = ", ".join(parts[1:])
        elif field == "NAME" and value:
            names.append(value)
        elif field == "PATHWAY" and value:
            pathways.append(value)
        elif field == "MODULE" and value:
            modules.append(value)
        elif field == "ORTHOLOGY" and value:
            orthology.append(value)
        for ec_block in re.findall(r"\[EC:([^\]]+)\]", value):
            enzymes.extend(f"ec:{ec}" for ec in re.split(r"\s+", ec_block.strip()) if ec)
    return {
        "symbol": symbol,
        "aliases": aliases,
        "name": " ".join(names),
        "pathways": ";".join(pathways),
        "modules": ";".join(modules),
        "orthology": ";".join(orthology),
        "enzymes": ";".join(unique(enzymes)),
    }


def parse_kegg_entry_field(text: str, wanted: str) -> str:
    values: list[str] = []
    current_field = ""
    for line in text.splitlines():
        field = line[:12].strip()
        value = line[12:].strip() if len(line) > 12 else ""
        if field:
            current_field = field
        else:
            field = current_field
        if field == wanted and value:
            values.append(value.rstrip(";"))
    return " ".join(values)


def parse_kegg_entry_gene_id(text: str) -> str:
    for line in text.splitlines():
        if not line.startswith("ENTRY"):
            continue
        value = line[12:].strip()
        if not value:
            continue
        gene_number = value.split()[0]
        return f"mmu:{gene_number}"
    return ""


def infer_mechanism_tags(gene_info: dict[str, str], pathway_name: str = "") -> str:
    text = " ".join(
        [
            gene_info.get("symbol", ""),
            gene_info.get("aliases", ""),
            gene_info.get("name", ""),
            gene_info.get("pathways", ""),
            gene_info.get("modules", ""),
            gene_info.get("orthology", ""),
            pathway_name,
        ]
    ).lower()
    tags: list[str] = []
    rules = [
        (("adenosylmethionine", "decarboxylase"), "polyamine metabolism"),
        (("s-adenosylmethionine",), "S-adenosylmethionine metabolism"),
        (("methionine salvage",), "methionine salvage"),
        (("cysteine and methionine",), "cysteine/methionine metabolism"),
        (("arginine and proline",), "arginine/proline metabolism"),
        (("glutathione",), "glutathione/redox metabolism"),
        (("one carbon",), "one-carbon metabolism"),
        (("folate",), "one-carbon metabolism"),
        (("tryptophan",), "tryptophan metabolism"),
        (("retinol",), "retinoid metabolism"),
        (("glycolysis",), "glycolysis"),
        (("oxidative phosphorylation",), "oxidative phosphorylation"),
        (("fatty acid",), "fatty-acid metabolism"),
        (("purine",), "purine metabolism"),
        (("pyrimidine",), "pyrimidine metabolism"),
    ]
    for required_terms, tag in rules:
        if all(term in text for term in required_terms):
            tags.append(tag)
    return ";".join(unique(tags))


def numeric_values(row: tuple[Any, ...], indexes: list[int]) -> list[float]:
    values: list[float] = []
    for index in indexes:
        value = safe_float(row[index]) if index < len(row) else None
        if value is not None:
            values.append(value)
    return values


def safe_float(value: Any) -> float | None:
    if value is None:
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if math.isnan(number) or math.isinf(number):
        return None
    return number


def safe_log2_ratio(numerator: float, denominator: float) -> float:
    return math.log2((numerator + 1e-9) / (denominator + 1e-9))


def clean_name(value: Any) -> str:
    return re.sub(r"\s+", " ", "" if value is None else str(value)).strip()


def normalize_metabolite_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", value.lower())


def is_kegg_friendly_metabolite(value: str) -> bool:
    lowered = value.lower()
    if any(token in lowered for token in ("ceramide", "isomer")):
        return False
    if re.match(r"^(pc|pe|pg|pi|ps|sm|dag|tag|tg)\s*[\(\[]", value, flags=re.IGNORECASE):
        return False
    if "/" in value:
        return False
    return bool(re.search(r"[a-zA-Z]", value))


def unique(values: list[str]) -> list[str]:
    seen: set[str] = set()
    output: list[str] = []
    for value in values:
        if value and value not in seen:
            seen.add(value)
            output.append(value)
    return output


def format_float(value: float | None | str) -> str:
    if value is None or value == "":
        return ""
    try:
        return f"{float(value):.3g}"
    except (TypeError, ValueError):
        return str(value)


def md_escape(value: str) -> str:
    return value.replace("|", "\\|")


def write_differentials_csv(path: Path, rows: list[DifferentialMetabolite]) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "source",
                "metabolite",
                "comparison",
                "direction",
                "log2fc",
                "control_mean",
                "disease_mean",
                "p_value",
                "fdr",
                "priority_score",
            ],
        )
        writer.writeheader()
        for row in rows:
            writer.writerow(
                {
                    "source": row.source,
                    "metabolite": row.metabolite,
                    "comparison": row.comparison,
                    "direction": row.direction,
                    "log2fc": row.log2fc,
                    "control_mean": row.control_mean,
                    "disease_mean": row.disease_mean,
                    "p_value": row.p_value,
                    "fdr": row.fdr,
                    "priority_score": row.priority_score,
                }
            )


def write_kegg_csv(path: Path, mappings: dict[str, KeggMapping]) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "metabolite",
                "status",
                "compound_ids",
                "compound_names",
                "pathways",
                "enzymes",
                "mouse_gene_symbols",
                "neighbor_mouse_gene_symbols",
            ],
        )
        writer.writeheader()
        for mapping in mappings.values():
            writer.writerow(
                {
                    "metabolite": mapping.metabolite,
                    "status": mapping.status,
                    "compound_ids": ";".join(mapping.compound_ids),
                    "compound_names": ";".join(mapping.compound_names),
                    "pathways": ";".join(mapping.pathways),
                    "enzymes": ";".join(mapping.enzymes),
                    "mouse_gene_symbols": ";".join(gene.get("symbol", "") for gene in mapping.mouse_genes),
                    "neighbor_mouse_gene_symbols": ";".join(
                        gene.get("symbol", "") for gene in mapping.neighbor_mouse_genes
                    ),
                }
            )


def write_gene_csv(path: Path, rows: list[dict[str, str]]) -> None:
    fieldnames = [
        "metabolite",
        "source",
        "comparison",
        "metabolite_direction",
        "metabolite_log2fc",
        "metabolite_priority_score",
        "compound_ids",
        "pathways",
        "link_type",
        "link_pathway",
        "link_pathway_name",
        "enzyme",
        "gene_id",
        "gene_symbol",
        "gene_aliases",
        "gene_name",
        "gene_pathways",
        "gene_modules",
        "mechanism_tags",
        "evidence_priority_score",
        "hypothesis_readiness_score",
    ]
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def write_pubmed_csv(path: Path, rows: list[dict[str, str]]) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "gene_symbol",
                "pmid",
                "title",
                "journal",
                "pubdate",
                "query_terms",
                "hit_type",
                "bridge_candidate",
                "evidence_role",
            ],
            extrasaction="ignore",
        )
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def read_csv_dicts(path: Path) -> list[dict[str, str]]:
    if not path.exists() or path.stat().st_size == 0:
        return []
    with path.open("r", newline="", encoding="utf-8") as handle:
        return list(csv.DictReader(handle))


def load_json(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def save_json(path: Path, data: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")


def mapping_to_dict(mapping: KeggMapping) -> dict[str, Any]:
    return {
        "schema_version": PATHWAY_NEIGHBOR_SCHEMA_VERSION,
        "compound_ids": mapping.compound_ids,
        "compound_names": mapping.compound_names,
        "pathways": mapping.pathways,
        "enzymes": mapping.enzymes,
        "mouse_genes": mapping.mouse_genes,
        "neighbor_mouse_genes": mapping.neighbor_mouse_genes,
        "neighbor_expanded": mapping.neighbor_expanded,
        "status": mapping.status,
    }


def mapping_from_dict(metabolite: str, data: dict[str, Any]) -> KeggMapping:
    return KeggMapping(
        metabolite=metabolite,
        compound_ids=list(data.get("compound_ids", [])),
        compound_names=list(data.get("compound_names", [])),
        pathways=list(data.get("pathways", [])),
        enzymes=list(data.get("enzymes", [])),
        mouse_genes=list(data.get("mouse_genes", [])),
        neighbor_mouse_genes=list(data.get("neighbor_mouse_genes", [])),
        neighbor_expanded=bool(data.get("neighbor_expanded")),
        status=str(data.get("status", "cached")),
    )
